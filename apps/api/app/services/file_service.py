from __future__ import annotations

import shutil
import urllib.request
import uuid
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import HTTPException, UploadFile

from app.core.config import ALLOWED_EXTENSIONS, HISTORY_DIR, UPLOAD_DIR
from app.core.storage import append_record, now_iso, read_json, write_json
from app.services.field_detector import detect_mapping
from app.services.object_storage import download_to_temp, presigned_download_url, r2_enabled


UPLOAD_INDEX = HISTORY_DIR / "uploads.json"


def read_dataframe(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, dtype=str, keep_default_na=False)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path, dtype=str, keep_default_na=False)
    raise HTTPException(status_code=400, detail="仅支持 xlsx、xls、csv 文件")


def read_dataframe_sample(path: Path, limit: int = 50) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, dtype=str, keep_default_na=False, nrows=limit)
    if suffix == ".xlsx":
        return _read_xlsx_sample(path, limit)
    if suffix == ".xls":
        return pd.read_excel(path, dtype=str, keep_default_na=False, nrows=limit)
    raise HTTPException(status_code=400, detail="仅支持 xlsx、xls、csv 文件")


def fetch_remote_file(source_url: str, filename: str | None = None) -> Path:
    suffix = Path(filename or source_url).suffix.lower() or ".xlsx"
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="远程文件格式不受支持")
    temp_path = UPLOAD_DIR / f"remote-{uuid.uuid4()}{suffix}"
    try:
        with urllib.request.urlopen(source_url, timeout=30) as response, temp_path.open("wb") as file:
            shutil.copyfileobj(response, file)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"远程文件读取失败: {exc}") from exc
    return temp_path


def _json_safe_rows(df: pd.DataFrame, limit: int = 20) -> list[dict[str, Any]]:
    rows = df.head(limit).where(pd.notna(df), "").to_dict(orient="records")
    return [{str(k): ("" if v is None else v) for k, v in row.items()} for row in rows]


def _read_xlsx_sample(path: Path, limit: int = 50) -> pd.DataFrame:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        buffered_rows: list[list[Any]] = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = ["" if value is None else value for value in row]
            if any(str(value).strip() for value in values):
                buffered_rows.append(values)
            if index >= limit + 10 or len(buffered_rows) >= limit + 1:
                break
    finally:
        workbook.close()

    if not buffered_rows:
        return pd.DataFrame()

    header_index = _detect_header_row(buffered_rows)
    headers = _unique_columns(buffered_rows[header_index])
    data_rows = buffered_rows[header_index + 1 : header_index + 1 + limit]
    normalized_rows = [row[: len(headers)] + [""] * max(len(headers) - len(row), 0) for row in data_rows]
    return pd.DataFrame(normalized_rows, columns=headers, dtype=str)


def _detect_header_row(rows: list[list[Any]]) -> int:
    search_rows = rows[: min(len(rows), 10)]
    scored = [
        (
            sum(1 for value in row if str(value).strip()),
            index,
        )
        for index, row in enumerate(search_rows)
    ]
    scored.sort(key=lambda item: (-item[0], item[1]))
    return scored[0][1] if scored and scored[0][0] > 0 else 0


def _unique_columns(values: list[Any]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        column = str(value).strip() if value is not None else ""
        if not column:
            column = f"列{index}"
        count = seen.get(column, 0)
        seen[column] = count + 1
        result.append(column if count == 0 else f"{column}_{count + 1}")
    return result


async def save_upload(file: UploadFile) -> dict[str, Any]:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持 xlsx、xls、csv 文件")

    upload_id = str(uuid.uuid4())
    stored_path = UPLOAD_DIR / f"{upload_id}{suffix}"
    with stored_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    df = read_dataframe(stored_path)
    columns = [str(column) for column in df.columns]
    sample_rows = _json_safe_rows(df)
    suggested_mapping = detect_mapping(columns, sample_rows)

    record = {
        "id": upload_id,
        "original_name": file.filename,
        "stored_path": str(stored_path),
        "file_type": suffix.lstrip("."),
        "rows_count": int(len(df)),
        "columns": columns,
        "sample_rows": sample_rows,
        "suggested_mapping": suggested_mapping,
        "created_at": now_iso(),
    }
    write_json(HISTORY_DIR / f"upload-{upload_id}.json", record)
    append_record(UPLOAD_INDEX, record)
    return record


def register_remote_upload(
    object_key: str,
    filename: str,
    content_type: str | None = None,
    columns: list[str] | None = None,
    sample_rows: list[dict[str, Any]] | None = None,
    rows_count: int | None = None,
) -> dict[str, Any]:
    if not r2_enabled():
        raise HTTPException(status_code=400, detail="对象存储未启用")
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持 xlsx、xls、csv 文件")

    upload_id = str(uuid.uuid4())
    if columns and sample_rows is not None:
        resolved_columns = [str(column) for column in columns]
        resolved_sample_rows = sample_rows[:20]
        resolved_rows_count = int(rows_count or len(sample_rows))
    else:
        try:
            source_path = download_to_temp(object_key, filename)
            df = read_dataframe_sample(source_path)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"对象存储文件读取失败: {exc}") from exc
        resolved_columns = [str(column) for column in df.columns]
        resolved_sample_rows = _json_safe_rows(df)
        resolved_rows_count = int(len(df))
    suggested_mapping = detect_mapping(resolved_columns, resolved_sample_rows)
    record = {
        "id": upload_id,
        "original_name": filename,
        "stored_path": "",
        "file_type": suffix.lstrip("."),
        "rows_count": resolved_rows_count,
        "columns": resolved_columns,
        "sample_rows": resolved_sample_rows,
        "suggested_mapping": suggested_mapping,
        "created_at": now_iso(),
        "storage_mode": "r2",
        "storage_key": object_key,
        "source_url": presigned_download_url(object_key, filename),
        "content_type": content_type,
    }
    write_json(HISTORY_DIR / f"upload-{upload_id}.json", record)
    append_record(UPLOAD_INDEX, record)
    return record


async def save_runtime_upload(file: UploadFile) -> Path:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持 xlsx、xls、csv 文件")
    stored_path = UPLOAD_DIR / f"runtime-{uuid.uuid4()}{suffix}"
    with stored_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return stored_path


def get_upload(upload_id: str) -> dict[str, Any]:
    record = read_json(HISTORY_DIR / f"upload-{upload_id}.json", None)
    if not record:
        raise HTTPException(status_code=404, detail="上传记录不存在")
    return record


def resolve_source(upload_id: str | None = None, source_url: str | None = None, source_filename: str | None = None) -> tuple[Path, str]:
    if upload_id:
        record = get_upload(upload_id)
        if record.get("storage_key") and r2_enabled():
            return download_to_temp(record["storage_key"], record["original_name"]), record["original_name"]
        return Path(record["stored_path"]), record["original_name"]
    if source_url:
        return fetch_remote_file(source_url, source_filename), source_filename or Path(source_url).name
    raise HTTPException(status_code=400, detail="缺少 upload_id 或 source_url")
