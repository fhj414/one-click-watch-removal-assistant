from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


MONEY_HEADERS = {"金额", "净额", "收入", "支出", "收入金额", "支出金额", "数值"}
DATE_HEADERS = {"日期"}


def export_workbook(tables: dict[str, pd.DataFrame], output_path: Path, enable_formulas: bool = True) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, df in tables.items():
        ws = wb.create_sheet(sheet_name[:31])
        safe_df = _prepare_for_excel(df)
        headers = [str(column) for column in safe_df.columns]
        ws.append(headers)
        for row in safe_df.itertuples(index=False, name=None):
            ws.append(list(row))

        _style_sheet(ws, headers, sheet_name)
        if enable_formulas and ws.max_row >= 2:
            _add_totals(ws, headers)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _prepare_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for column in result.columns:
        if pd.api.types.is_datetime64_any_dtype(result[column]):
            result[column] = result[column].dt.date
    return result.where(pd.notna(result), "")


def _style_sheet(ws, headers: list[str], sheet_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="E8F1FF")
    anomaly_fill = PatternFill("solid", fgColor="FFE4E6")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=_safe_table_name(sheet_name), ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        width = min(max(len(header) + 6, 12), 28)
        ws.column_dimensions[letter].width = width
        if header in MONEY_HEADERS:
            for cell in ws[letter][1:]:
                cell.number_format = '#,##0.00'
        if header in DATE_HEADERS:
            for cell in ws[letter][1:]:
                cell.number_format = "yyyy-mm-dd"

    if "异常原因" in headers and ws.max_row >= 2:
        reason_col = get_column_letter(headers.index("异常原因") + 1)
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ws.max_column)}{ws.max_row}",
            FormulaRule(formula=[f'LEN(${reason_col}2)>0'], fill=anomaly_fill),
        )


def _add_totals(ws, headers: list[str]) -> None:
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    for index, header in enumerate(headers, start=1):
        if header in MONEY_HEADERS:
            letter = get_column_letter(index)
            ws.cell(row=total_row, column=index, value=f"=SUBTOTAL(109,{letter}2:{letter}{total_row - 1})")
            ws.cell(row=total_row, column=index).number_format = '#,##0.00'
            ws.cell(row=total_row, column=index).font = Font(bold=True)


def _safe_table_name(sheet_name: str) -> str:
    ascii_name = "".join(ch for ch in sheet_name if ch.isascii() and ch.isalnum())
    return f"T_{ascii_name or abs(hash(sheet_name))}"
